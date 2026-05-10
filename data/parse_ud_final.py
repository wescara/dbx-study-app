import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

try:
    from ai_explanations import EXPLANATIONS
except ImportError:
    EXPLANATIONS = {}

def generate_explanation_with_ai(qid, question_text, correct_answer, options):
    """Get explanation from AI-generated explanations or generate fallback."""
    # Check if we have a pre-generated explanation for this QID
    if qid in EXPLANATIONS:
        return EXPLANATIONS[qid]
    
    # Otherwise use fallback
    return generate_explanation_fallback(question_text, correct_answer, options)

def generate_explanation_fallback(question_text, correct_answer, options):
    """Fallback explanation generation when AI is not available."""
    correct_answer_text = options.get(correct_answer, "")
    q_lower = question_text.lower()
    
    # Generic template-based explanations
    if 'which' in q_lower and 'not' in q_lower:
        return f"The correct answer is {correct_answer}. {correct_answer_text} is the only option that does NOT match the criteria described in the question."
    
    if 'which' in q_lower and 'best' in q_lower:
        return f"The correct answer is {correct_answer}. {correct_answer_text} is the best approach because it aligns with Databricks best practices."
    
    if 'which' in q_lower and 'can' in q_lower:
        return f"The correct answer is {correct_answer}. Among the options, only {correct_answer_text} fulfills the requirement described in the question."
    
    if 'fill in' in q_lower or 'blank' in q_lower:
        return f"The correct answer is {correct_answer}. The blank should be filled with '{correct_answer_text}' to make the syntax correct."
    
    if 'error' in q_lower or 'invalid' in q_lower or 'wrong' in q_lower:
        return f"The correct answer is {correct_answer}. The issue is: {correct_answer_text}. This represents the correct fix or identification of the problem."
    
    if 'true' in q_lower or 'false' in q_lower or 'not true' in q_lower:
        return f"The correct answer is {correct_answer}. {correct_answer_text} is the statement that best reflects Databricks functionality."
    
    if 'how' in q_lower or 'what' in q_lower:
        return f"The correct answer is {correct_answer}. {correct_answer_text} is the correct response to the question based on Databricks documentation and best practices."
    
    if 'default' in q_lower:
        return f"The correct answer is {correct_answer}. By default, {correct_answer_text}."
    
    # Fallback explanation
    return f"The correct answer is {correct_answer}. {correct_answer_text}"


# Read the entire file
with open('UDQuestionsAnswers.txt', 'r', encoding='utf-8') as f:
    content = f.read()

# Split by the QID pattern (lines starting with -)
qid_pattern = r'^-(\d+)-+'
sections = re.split(qid_pattern, content, flags=re.MULTILINE)

if EXPLANATIONS:
    print(f"✓ Loaded {len(EXPLANATIONS)} AI-generated explanations")
else:
    print("ℹ Using fallback explanations")

# sections will be: ['', 'QID', 'content', 'QID', 'content', ...]
questions = []

for i in range(1, len(sections), 2):
    if i + 1 >= len(sections):
        break
    
    qid = int(sections[i].strip())
    section_content = sections[i + 1].strip()
    
    # Split the section into lines
    lines = [l for l in section_content.split('\n')]
    
    # First non-empty line should be the question
    question_text = ""
    options = {}
    correct_answer = None
    explanation = ""
    
    line_idx = 0
    
    # Get the question (multi-line support)
    while line_idx < len(lines) and not lines[line_idx].strip():
        line_idx += 1
    
    if line_idx < len(lines):
        question_text = lines[line_idx].strip()
        line_idx += 1
        
        # Continue reading question lines until we hit an option or explanation
        while line_idx < len(lines):
            line = lines[line_idx].strip()
            
            # Stop if we hit an option (starts with A., B., C., etc.)
            if re.match(r'^[A-F]\.\s+', line):
                break
            
            # Stop if we hit explanation
            if line.startswith('Explanation:'):
                break
            
            # Add non-empty lines to question
            if line:
                question_text += '\n' + line
            
            line_idx += 1
    
    # Parse options - when we see a duplicate letter, that's the answer
    seen_letters = set()
    while line_idx < len(lines):
        line = lines[line_idx].strip()
        
        if not line:
            line_idx += 1
            continue
        
        # Stop if we hit explanation
        if line.startswith('Explanation:'):
            break
        
        # Check if line starts with option letter
        match = re.match(r'^([A-F])\.\s+(.*)', line)
        if match:
            letter = match.group(1)
            text = match.group(2).strip()
            
            # If we've seen this letter before, it's the answer
            if letter in seen_letters:
                correct_answer = letter
                line_idx += 1
                break
            
            options[letter] = text
            seen_letters.add(letter)
            line_idx += 1
        else:
            # Not an option line, skip
            line_idx += 1
    
    # Skip to explanation
    while line_idx < len(lines) and not lines[line_idx].strip().startswith('Explanation:'):
        line_idx += 1
    
    # Get explanation (multi-line support)
    if line_idx < len(lines):
        line = lines[line_idx].strip()
        if line.startswith('Explanation:'):
            explanation = line.replace('Explanation:', '').strip()
            line_idx += 1
            # Continue reading explanation lines until we hit empty line or end
            while line_idx < len(lines):
                next_line = lines[line_idx].strip()
                # Stop if we hit an empty line or end of section
                if not next_line:
                    break
                # Stop if we hit the next QID marker (starts with many dashes)
                if re.match(r'^-+$', next_line):
                    break
                # Add continuation line to explanation
                explanation += '\n' + next_line
                line_idx += 1
    
    # Validate we have minimum data
    if not question_text or len(options) < 4 or not correct_answer:
        continue
    
    # Generate explanation if missing or N/A
    if not explanation or explanation.lower() == 'n/a':
        explanation = generate_explanation_with_ai(qid, question_text, correct_answer, options)
    
    # Topic/Subtopic assignment
    q_lower = question_text.lower()
    
    if 'repos' in q_lower or 'git' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Repos and Git Operations'
    elif 'delta lake' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Delta Lake'
    elif 'vacuum' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Delta Lake Operations'
    elif 'auto loader' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Auto Loader'
    elif 'dlt' in q_lower or 'delta live table' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Delta Live Tables'
    elif 'pyspark' in q_lower or 'spark.sql' in q_lower or 'spark' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Spark SQL and PySpark'
    elif 'python' in q_lower or 'if-condition' in q_lower or 'if condition' in q_lower or 'syntax' in q_lower or 'code' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Python Syntax'
    elif 'lakehouse' in q_lower or 'architecture' in q_lower:
        topic, subtopic = 'Databricks Intelligence Platform', 'Lakehouse Architecture'
    elif 'structured streaming' in q_lower or 'trigger' in q_lower or 'streaming' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Structured Streaming'
    elif 'sql warehouse' in q_lower or 'databricks sql' in q_lower:
        topic, subtopic = 'Databricks Intelligence Platform', 'SQL Warehouse'
    elif 'job' in q_lower or 'task' in q_lower or 'orchestration' in q_lower:
        topic, subtopic = 'Databricks Intelligence Platform', 'Jobs and Orchestration'
    elif 'permission' in q_lower or 'grant' in q_lower or 'rbac' in q_lower:
        topic, subtopic = 'Databricks Intelligence Platform', 'Access Control'
    elif 'merge' in q_lower or 'insert' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'Delta Table Operations'
    elif 'jdbc' in q_lower or 'postgresql' in q_lower or 'database' in q_lower:
        topic, subtopic = 'Development and Ingestion', 'External Data Integration'
    else:
        topic, subtopic = 'Development and Ingestion', 'General'
    
    # Build question row
    row = {
        'Source': 'UD',
        'QID': qid,
        'Exam': 1,
        'Number': qid - 1769,
        'Topic': topic,
        'Subtopic': subtopic,
        'Difficulty': 'Medium',
        'Question': question_text,
        'A': options.get('A', ''),
        'B': options.get('B', ''),
        'C': options.get('C', ''),
        'D': options.get('D', ''),
        'E': options.get('E', ''),
        'F': options.get('F', ''),
        'CorrectLetter': correct_answer,
        'Explanation': explanation,
        'Notes': '',
        'LastReviewed': datetime.now().date(),
        'CorrectCount': 0,
        'IncorrectCount': 0,
        'FlagForReview': False,
        'Tags': '',
        'Graphic': '',
        'Deeper Dive': '',
        'VerifiedAnswer': correct_answer,
        'VerificationStatus': 'Confirmed',
        'AnswerMatchesWorkbook': '',
        'VerificationNotes': 'Custom question from user study materials',
        'VerificationSourceKeys': '',
        'VerifiedOn': '',
    }
    
    questions.append(row)

# Create DataFrame and save
df = pd.DataFrame(questions)
if len(df) > 0:
    output_file = 'Custom_Questions_All_Test.xlsx'
    temp_file = 'Custom_Questions_All_Test_temp.xlsx'
    
    # Remove temp file if it exists
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except:
            pass
    
    # Write to temp file first
    try:
        df.to_excel(temp_file, sheet_name='Custom Questions', index=False)
        
        # Apply text wrapping to Question and Explanation columns
        wb = load_workbook(temp_file)
        ws = wb.active
        
        # Find column indices for Question and Explanation
        headers = [cell.value for cell in ws[1]]
        question_col = headers.index('Question') + 1 if 'Question' in headers else None
        explanation_col = headers.index('Explanation') + 1 if 'Explanation' in headers else None
        
        # Apply wrapping to these columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if question_col:
                row[question_col - 1].alignment = Alignment(wrap_text=True, vertical='top')
            if explanation_col:
                row[explanation_col - 1].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Set column widths for readability using column letters
        if question_col:
            ws.column_dimensions[get_column_letter(question_col)].width = 50
        if explanation_col:
            ws.column_dimensions[get_column_letter(explanation_col)].width = 50
        
        wb.save(temp_file)
        
        # Replace original file
        if os.path.exists(output_file):
            try:
                os.remove(output_file)
            except PermissionError:
                print(f"Warning: Could not remove {output_file} (file is in use). Saving as {temp_file}")
                output_file = temp_file
                temp_file = None
        
        if temp_file:
            os.rename(temp_file, output_file)
        
        print(f"✓ Created test workbook: {output_file}")
        print(f"✓ Added {len(df)} questions")
        if len(df) > 0:
            print(f"✓ QID range: {df['QID'].min()} to {df['QID'].max()}")
        print(f"\nQuestions by topic:")
        for topic in sorted(df['Topic'].unique()):
            count = len(df[df['Topic'] == topic])
            print(f"  - {topic}: {count} questions")
    except Exception as e:
        print(f"ERROR: Could not save Excel file: {e}")
else:
    print("ERROR: No questions parsed!")
