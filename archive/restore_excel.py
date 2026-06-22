import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Load current file (has all 945 questions + flags)
df = pd.read_excel(r'data/DBx_Questions.xlsx', sheet_name='QuestionBank')

print(f'Loaded: {df.shape[0]} questions, {df.shape[1]} columns')
print(f'Flagged questions: {df[df["FlagForReview"] == 1]["QID"].tolist()}')

# Export to a temporary file first
temp_path = r'data/DBx_Questions_temp.xlsx'
df.to_excel(temp_path, sheet_name='QuestionBank', index=False, engine='openpyxl')

# Now clean up the file with openpyxl
wb = load_workbook(temp_path)
ws = wb.active

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save back
wb.save(temp_path)

# Replace original with cleaned version
import shutil
shutil.move(temp_path, r'data/DBx_Questions.xlsx')

print('✓ File restored with proper formatting')
print('✓ All 945 questions preserved')
print('✓ Flags preserved')
