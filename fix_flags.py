import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

df = pd.read_excel(r'data/DBx_Questions.xlsx', sheet_name='QuestionBank')

# Clear all flags first
df['FlagForReview'] = 0

# Flag ONLY the 5 missed ones
qids_to_flag = [2120, 2121, 2124, 2131, 2132]
df.loc[df['QID'].isin(qids_to_flag), 'FlagForReview'] = 1

# Count flags
flag_count = (df['FlagForReview'] == 1).sum()
print(f'Total flagged: {flag_count}')
print(f'Flagged QIDs: {sorted(df[df["FlagForReview"] == 1]["QID"].tolist())}')

# Save with openpyxl to preserve formatting
temp_path = r'data/DBx_Questions_temp.xlsx'
df.to_excel(temp_path, sheet_name='QuestionBank', index=False, engine='openpyxl')

wb = load_workbook(temp_path)
ws = wb.active

for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

wb.save(temp_path)

import shutil
shutil.move(temp_path, r'data/DBx_Questions.xlsx')

print('✓ Fixed - only 5 questions flagged')
